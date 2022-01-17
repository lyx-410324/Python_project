#!/usr/bin/env python
# @Time:
# @Author:LYX
# @File:exclefunction.py.py
# @Software:PyCharm
import os
import time
#调用库函数
from openpyxl import Workbook,load_workbook
#利用这些可以自行设置字体和颜色
from openpyxl.styles import Font,colors
import win32api,win32con
EXCELPATH = './果谷智能锁注册记录.xlsx'
SHEETNAME = '果谷智能锁注册记录'
MACTXTDICTORYPATH = './Recorder/'

def write_excel_xlsx1(path, sheet_name, value):
    #index = len(value)#列表中所含元组的个数，从而确定写入Excel的行数

    #打开Excel
    wb=load_workbook(path)
    #创建工作簿
    #sheet=wb.create_sheet(sheet_name)
    sheet = wb[sheet_name] #wb.get_sheet_by_name(sheet_name)#弃用wb.get_sheet_by_name(sheet_name)
    #设置格式
    sheet.column_dimensions['A'].width=40
    sheet.column_dimensions['B'].width=30
    sheet.column_dimensions['C'].width=30
    sheet.append(value)
    #保存文件
    wb.save(path)
    print("题目写入成功！")
def CreateExcel():
    if not os.path.exists(EXCELPATH):
        book_name_xlsx =  EXCELPATH #文件路径，根据自己的需要自行修改
        wb = Workbook()
        sheet_name_xlsx = SHEETNAME
        wb.create_sheet(SHEETNAME)
        wb.save(book_name_xlsx)
        listinfo = ['门锁序列号','门锁试用码','门锁激活码','日期']
        write_excel_xlsx1(book_name_xlsx, sheet_name_xlsx,listinfo)#调用写入函数
    else:
        print('文件已存在')
def InsertExcel(path, sheet_name, value):
    value.insert(3,getDate())
    print(value)
    if not os.path.exists(path):
        CreateExcel()
    # 打开文件
    win32api.SetFileAttributes(path, win32con.FILE_ATTRIBUTE_NORMAL)
    wb = load_workbook(path)
    # 打开工作簿
    sheet = wb[sheet_name] #wb.get_sheet_by_name(sheet_name)#弃用
    sheet.append(value)
    # 保存文件
    wb.save(path)
    win32api.SetFileAttributes(path, win32con.FILE_ATTRIBUTE_READONLY)
    print("写入数据成功！")
def SaveMacAndActivationCode(value):
    InsertExcel(EXCELPATH, SHEETNAME, value)

#取得当前时间戳
def getDate():
    return str(time.strftime('%Y.%m.%d %H:%M:%S', time.localtime(time.time())))

#保存序列号文本

def SaveMacInDictory(MacCode,TryoutCode):
    path = MACTXTDICTORYPATH + str(MacCode)+'.dat'
    content = '设备MAC:'+str(MacCode)+';'+'试用激活码:'+str(TryoutCode)
    fd = open(path,'w',encoding='utf-8')
    fd.write(content)
    win32api.SetFileAttributes(path, win32con.FILE_ATTRIBUTE_READONLY)
    fd.close()


#格式化时间戳为标准格式
if __name__ == '__main__':
    #CreateExcel()
    InsertExcel(EXCELPATH,SHEETNAME,[12,12,12])